"""
sentinel/core/ingest_extended.py
---------------------------------
Multi-format ingestion layer for Sentinel.

Supports:
  .json  — structured JSON (existing format or generic array/object)
  .csv   — comma / tab separated values
  .xml   — any XML; heuristic element extraction
  .xls / .xlsx — Excel workbooks (any sheet)
  .pdf   — text extraction + Gemini Vision fallback for scanned/image PDFs
  .png / .jpg / .jpeg / .webp / .bmp — image documents via Gemini Vision

Each parser returns a ParseResult containing:
  - normalized records (list[dict])
  - detected document_type: "invoice" | "purchase_order" | "delivery" | "unknown"
  - SourceRef provenance for downstream evidence chain

External deps (install as needed):
  pip install openpyxl xlrd pdfplumber google-generativeai pillow
"""

from __future__ import annotations

import base64
import io
import json
import logging
import csv
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

# ── Result types ──────────────────────────────────────────────────────────────

@dataclass
class ParseResult:
    file_path: str
    file_type: str                     # extension
    document_type: str                 # invoice | purchase_order | delivery | unknown
    records: list[dict] = field(default_factory=list)
    raw_text: str = ""                 # for audit / debug
    errors: list[str] = field(default_factory=list)
    page_count: int = 0
    used_vision: bool = False          # True when Gemini Vision was invoked


# ── Document-type heuristic ───────────────────────────────────────────────────

_INV_SIGNALS  = {"invoice", "inv", "bill", "billed", "vendor", "billed_unit_price"}
_PO_SIGNALS   = {"purchase", "po", "order", "agreed_unit_price", "qty_authorized", "authorized"}
_POD_SIGNALS  = {"delivery", "pod", "waybill", "received", "dock", "condition", "waybill_ref"}


def _detect_doc_type(keys: list[str]) -> str:
    key_set = {k.lower() for k in keys}
    def score(signals): return sum(1 for s in signals if any(s in k for k in key_set))
    scores = {"invoice": score(_INV_SIGNALS), "purchase_order": score(_PO_SIGNALS), "delivery": score(_POD_SIGNALS)}
    best = max(scores, key=scores.get)
    return best if scores[best] > 0 else "unknown"


# ── JSON ──────────────────────────────────────────────────────────────────────

def parse_json(path: Path) -> ParseResult:
    result = ParseResult(file_path=str(path), file_type="json", document_type="unknown")
    try:
        raw = path.read_text(encoding="utf-8")
        data = json.loads(raw)
        if isinstance(data, list):
            records = data
        elif isinstance(data, dict):
            # Try to unwrap envelope key (e.g. {"invoices": [...]} or {"value": [...]})
            for v in data.values():
                if isinstance(v, list):
                    records = v
                    break
            else:
                records = [data]
        else:
            records = [{"value": data}]
        result.records = records
        all_keys = list(records[0].keys()) if records else []
        for _r in records[:3]:
            for _v in _r.values() if isinstance(_r, dict) else []:
                if isinstance(_v, list) and _v and isinstance(_v[0], dict):
                    all_keys.extend(_v[0].keys())
                elif isinstance(_v, dict):
                    all_keys.extend(_v.keys())
        result.document_type = _detect_doc_type(all_keys)
    except Exception as exc:
        result.errors.append(str(exc))
    return result


# ── CSV ───────────────────────────────────────────────────────────────────────

def parse_csv(path: Path) -> ParseResult:
    result = ParseResult(file_path=str(path), file_type="csv", document_type="unknown")
    try:
        records = []
        with path.open(newline="", encoding="utf-8-sig") as fh:
            reader = csv.DictReader(fh)
            for row in reader:
                records.append(dict(row))
        result.records = records
        result.document_type = _detect_doc_type(list(records[0].keys()) if records else [])
    except Exception as exc:
        result.errors.append(str(exc))
    return result


# ── XML ───────────────────────────────────────────────────────────────────────

def _elem_to_dict(elem) -> dict:
    """Recursively flatten an XML element into a dict."""
    d = dict(elem.attrib)
    for child in elem:
        tag = child.tag.split("}")[-1]  # strip namespace
        child_dict = _elem_to_dict(child)
        if tag in d:
            if not isinstance(d[tag], list):
                d[tag] = [d[tag]]
            d[tag].append(child_dict)
        else:
            d[tag] = child_dict if child_dict else (child.text or "").strip()
    if not d and elem.text:
        return {"_text": elem.text.strip()}
    return d


def parse_xml(path: Path) -> ParseResult:
    result = ParseResult(file_path=str(path), file_type="xml", document_type="unknown")
    try:
        tree = ET.parse(str(path))
        root = tree.getroot()
        # If root has children, treat each direct child as a record
        children = list(root)
        if children:
            records = [_elem_to_dict(child) for child in children]
        else:
            records = [_elem_to_dict(root)]
        result.records = records
        all_keys: list[str] = []
        for r in records:
            all_keys.extend(r.keys())
            if isinstance(r, dict):
                for v in r.values():
                    if isinstance(v, dict):
                        all_keys.extend(v.keys())
        result.document_type = _detect_doc_type(all_keys)
    except Exception as exc:
        result.errors.append(str(exc))
    return result


# ── Excel (XLS / XLSX) ────────────────────────────────────────────────────────

def parse_excel(path: Path) -> ParseResult:
    ext = path.suffix.lower()
    result = ParseResult(file_path=str(path), file_type=ext.lstrip("."), document_type="unknown")
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        all_records = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            # First non-empty row = headers
            headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
            for row in rows[1:]:
                if any(cell is not None for cell in row):
                    record = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
                    record["_sheet"] = sheet_name
                    all_records.append(record)
        result.records = all_records
        result.document_type = _detect_doc_type(list(all_records[0].keys()) if all_records else [])
    except ImportError:
        # Fallback to xlrd for .xls
        try:
            import xlrd
            wb = xlrd.open_workbook(str(path))
            all_records = []
            for sheet in wb.sheets():
                if sheet.nrows < 2:
                    continue
                headers = [str(sheet.cell_value(0, c)).strip() or f"col_{c}"
                           for c in range(sheet.ncols)]
                for r in range(1, sheet.nrows):
                    record = {headers[c]: sheet.cell_value(r, c) for c in range(sheet.ncols)}
                    record["_sheet"] = sheet.name
                    all_records.append(record)
            result.records = all_records
            result.document_type = _detect_doc_type(list(all_records[0].keys()) if all_records else [])
        except Exception as exc2:
            result.errors.append(str(exc2))
    except Exception as exc:
        result.errors.append(str(exc))
    return result


# ── PDF ───────────────────────────────────────────────────────────────────────

def _pdf_text_extract(path: Path) -> tuple[str, int]:
    """Extract raw text using pdfplumber. Returns (text, page_count)."""
    try:
        import pdfplumber
        pages = []
        with pdfplumber.open(str(path)) as pdf:
            for pg in pdf.pages:
                pages.append(pg.extract_text() or "")
        return "\n".join(pages), len(pages)
    except ImportError:
        raise RuntimeError("pdfplumber not installed: pip install pdfplumber")


def _pdf_to_images_base64(path: Path) -> list[str]:
    """Convert PDF pages to base64 PNG strings for vision processing."""
    try:
        from pdf2image import convert_from_path
        images = convert_from_path(str(path), dpi=150)
        out = []
        for img in images:
            buf = io.BytesIO()
            img.save(buf, format="PNG")
            out.append(base64.b64encode(buf.getvalue()).decode())
        return out
    except ImportError:
        raise RuntimeError("pdf2image not installed: pip install pdf2image")


def parse_pdf(path: Path, use_vision_fallback: bool = True) -> ParseResult:
    result = ParseResult(file_path=str(path), file_type="pdf", document_type="unknown")
    try:
        text, n_pages = _pdf_text_extract(path)
        result.page_count = n_pages
        result.raw_text = text

        # Heuristic: if we extracted meaningful text, try to parse as structured data
        if len(text.strip()) > 100:
            records, doc_type = _llm_extract_from_text(text, str(path))
            result.records = records
            result.document_type = doc_type
        elif use_vision_fallback:
            logger.info("PDF text extraction minimal — falling back to Vision for %s", path.name)
            result = parse_image_via_vision(path, result)
        else:
            result.errors.append("PDF text extraction returned minimal content.")
    except Exception as exc:
        result.errors.append(str(exc))
        if use_vision_fallback:
            try:
                result = parse_image_via_vision(path, result)
            except Exception as exc2:
                result.errors.append(f"Vision fallback also failed: {exc2}")
    return result


# ── Image (PNG / JPG) via Gemini Vision ──────────────────────────────────────

def _image_to_base64(path: Path) -> tuple[str, str]:
    """Returns (base64_data, media_type)."""
    ext_map = {".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
               ".webp": "image/webp", ".bmp": "image/bmp", ".gif": "image/gif"}
    media_type = ext_map.get(path.suffix.lower(), "image/png")
    data = base64.b64encode(path.read_bytes()).decode()
    return data, media_type


def parse_image_via_vision(
    path: Path,
    result: ParseResult | None = None,
) -> ParseResult:
    """Send an image (or PDF page) to Gemini Vision and extract structured records."""
    if result is None:
        result = ParseResult(
            file_path=str(path),
            file_type=path.suffix.lstrip(".").lower(),
            document_type="unknown",
        )
    result.used_vision = True

    try:
        from .llm import get_client
        client = get_client()
        if not client.available:
            result.errors.append("Gemini client unavailable — set GEMINI_API_KEY.")
            return result

        if path.suffix.lower() == ".pdf":
            b64_pages = _pdf_to_images_base64(path)
            if not b64_pages:
                result.errors.append("Could not convert PDF to image for Vision.")
                return result
            images = [{"data": b64_pages[0], "mime_type": "image/png"}]
        else:
            b64, media_type = _image_to_base64(path)
            images = [{"data": b64, "mime_type": media_type}]

        prompt = """You are a document data extraction assistant for a procurement reconciliation system.

Examine this document image carefully and extract ALL structured data.

Return ONLY a valid JSON object — no markdown fences, no commentary:
{
  "document_type": "<invoice|purchase_order|delivery_receipt|unknown>",
  "records": [ { ...one flat dict per line item... } ],
  "summary": "<one sentence description>"
}

Rules:
- Each line item is its own object in records.
- Use snake_case keys.
- Preserve numeric values as numbers.
- Return empty records array if nothing extractable.
"""
        parsed = client.complete_json(prompt, images=images, max_tokens=2000)
        result.records = parsed.get("records", [])
        result.document_type = parsed.get("document_type", "unknown")
        result.raw_text = parsed.get("summary", "")
    except Exception as exc:
        result.errors.append(f"Gemini Vision extraction failed: {exc}")

    return result


# ── LLM text-to-records (for text-heavy PDFs) ────────────────────────────────

def _llm_extract_from_text(text: str, source_label: str) -> tuple[list[dict], str]:
    """Ask Gemini to extract structured records from raw PDF text."""
    try:
        from .llm import get_client
        client = get_client()
        if not client.available:
            return [], "unknown"
        prompt = f"""Extract ALL structured line-item data from this procurement document text.

Return ONLY valid JSON (no markdown fences):
{{
  "document_type": "<invoice|purchase_order|delivery_receipt|unknown>",
  "records": [ {{ ...one dict per line item... }} ]
}}

Document text:
{text[:6000]}
"""
        parsed = client.complete_json(prompt, max_tokens=2000)
        return parsed.get("records", []), parsed.get("document_type", "unknown")
    except Exception as exc:
        logger.warning("Gemini text extraction failed: %s", exc)
        return [], "unknown"


# ── Universal dispatcher ──────────────────────────────────────────────────────

PARSERS: dict[str, Any] = {
    ".json": parse_json,
    ".csv":  parse_csv,
    ".xml":  parse_xml,
    ".xls":  parse_excel,
    ".xlsx": parse_excel,
    ".pdf":  parse_pdf,
    ".png":  parse_image_via_vision,
    ".jpg":  parse_image_via_vision,
    ".jpeg": parse_image_via_vision,
    ".webp": parse_image_via_vision,
    ".bmp":  parse_image_via_vision,
}


def parse_file(path: str | Path, use_vision: bool = True) -> ParseResult:
    """Parse a single file of any supported type. Returns a ParseResult."""
    path = Path(path)
    ext  = path.suffix.lower()

    if not path.exists():
        return ParseResult(
            file_path=str(path), file_type=ext.lstrip("."),
            document_type="unknown", errors=[f"File not found: {path}"]
        )

    parser = PARSERS.get(ext)
    if parser is None:
        return ParseResult(
            file_path=str(path), file_type=ext.lstrip("."),
            document_type="unknown",
            errors=[f"Unsupported file type: {ext}. Supported: {list(PARSERS.keys())}"]
        )

    logger.info("Parsing %s (%s) …", path.name, ext)
    if ext == ".pdf":
        return parser(path, use_vision_fallback=use_vision)
    elif ext in {".png", ".jpg", ".jpeg", ".webp", ".bmp"}:
        return parser(path)
    else:
        return parser(path)


def parse_files(paths: list[str | Path], use_vision: bool = True) -> list[ParseResult]:
    """Parse multiple files and return one ParseResult per file."""
    return [parse_file(p, use_vision=use_vision) for p in paths]
